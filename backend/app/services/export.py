"""Builds an Excel order sheet from a sourcing result.

The sheet is deliberately purchase-oriented: each row is a BOM line with the
chosen supplier and a direct product link you can click to add it to the cart,
plus the quantity and price so it doubles as a shopping list.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import SourcingResult

_HEADERS = [
    "Component",
    "Qty",
    "Supplier",
    "Unit Price (INR)",
    "Landed (INR)",
    "Link (add to cart)",
]
_WIDTHS = [34, 6, 18, 16, 14, 60]


def build_sourcing_workbook(result: SourcingResult) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Sheet"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    for col, (title, width) in enumerate(zip(_HEADERS, _WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    row_idx = 2
    for line in result.lines:
        component = line.matched_mpn or line.input.mpn or line.input.description or "Unknown part"
        qty = line.input.quantity
        if line.chosen is not None:
            offer = line.chosen
            ws.cell(row=row_idx, column=1, value=component)
            ws.cell(row=row_idx, column=2, value=qty)
            ws.cell(row=row_idx, column=3, value=offer.offer.distributor)
            ws.cell(row=row_idx, column=4, value=offer.unit_price_inr)
            ws.cell(row=row_idx, column=5, value=offer.landed_cost_inr)
            url = offer.offer.product_url or ""
            link_cell = ws.cell(row=row_idx, column=6, value=url or "-")
            if url:
                link_cell.hyperlink = url
                link_cell.font = Font(color="2563EB", underline="single")
        else:
            ws.cell(row=row_idx, column=1, value=component)
            ws.cell(row=row_idx, column=2, value=qty)
            ws.cell(row=row_idx, column=3, value="No supplier found")
        row_idx += 1

    # Totals row.
    row_idx += 1
    ws.cell(row=row_idx, column=3, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=row_idx, column=5, value=result.summary.total_landed_cost_inr)
    total_cell.font = Font(bold=True)

    return wb
